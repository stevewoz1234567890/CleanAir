import json
import sys
import os
import asyncio
import motor.motor_asyncio
from dateutil.parser import parse
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook, Workbook
import boto3
from statistics import mean
import re
from bson import ObjectId
from pathlib import Path  # erick
from collections import defaultdict  # seperate data by header
import copy
# import config as cfg

DEV = False  # specific for erick's env. Turn off with False


"""
TODO
- Add feature for events that extend beyond start and end times but fall within that timespan

- calculate values for 'to header' formulas

"""


def print_json(data):
    print(json.dumps(data, indent=4, default=str))


def _json(data):
    print(json.dumps(data, indent=4, default=str))


def lambda_handler(event=None, context=None):
    message = "successfully created report"
    response = {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Headers': 'Content-Type,X-Api-Key',
            'Access-Control-Allow-Origin': '*',
        }
    }
    print("event: ", event)
    try:
        link = asyncio.run(
            main(event)
        )
    except Exception as err:
        print("ERROR: ", err)
        response["statusCode"] = 500
        link = None
        message = "Error occurred. Talk to Dev Team"
    response["body"] = json.dumps(
        {"link": link, "message": message}, default=str)
    print(response)
    return response


class Mongo():
    def __init__(self, event, database_name=os.environ["db_name"]):
        self.database_name = database_name
        self.event = event
        self.conn_string = os.environ["conn_string"]
        self.client = motor.motor_asyncio.AsyncIOMotorClient(self.conn_string)

    def get_collection(self, coll_name):
        db = self.client[self.database_name]
        coll = db[coll_name]
        return coll


class WidgetMain():
    def __init__(self, event):
        self.mongo = Mongo(event)
        self.flare_id = event['flare_id']  # the flare id string
        self.flare = None  # Flare object
        self.rule_ids = event['rule_ids']  # the id of each report we want
        self.rules = None  # Rule objects
        self.events = {}
        self.s3_client = self.get_s3_client()
        self.bucket_name = 'flare-reporting'
        self.start_date = event['start_date']  # start date
        self.end_date = event['end_date']  # end date
        self.template_name = 'report_template.xlsx'
        self.report_resolution = "15 Minutes"
        self.headers = []

    async def setup(self):
        await self.get_parameters()
        await self.get_flare()
        await self.get_rules()
        await self.get_events()
        await self.get_formulas()
        await self.get_headers()
        await self.get_print_formulas()

    async def get_parameters(self):
        collection = self.mongo.get_collection(coll_name="parameters")
        self.parameters = [row for row in await collection.find({}).to_list(length=None)]

    async def get_flare(self):
        flares_collection = self.mongo.get_collection(coll_name="flares")
        self.flare = await flares_collection.find_one({"_id": ObjectId(self.flare_id)})

    async def get_rules(self):
        rules_collection = self.mongo.get_collection(coll_name="event_rules")
        rules = []
        for id in self.rule_ids:
            rule = await rules_collection.find_one({"_id": ObjectId(id)})
            if rule == None:
                print("No matching rule was found")
            else:
                rules.append(rule)
        self.rules = rules

    async def get_events(self):
        """Get the events for each rule"""
        async def get_rule_events(rule):
            """Get the events for a single rule"""
            mongo_filter = {
                'flare_id': ObjectId(self.flare_id),
                'rule_id': rule["_id"],
                'start': {
                    "$gte": parse(self.start_date) + timedelta(minutes=15)
                },
                'end': {
                    "$lte": parse(self.end_date) + timedelta(days=1)
                }
            }
            events = await collection.find(mongo_filter).to_list(length=None)
            self.events[str(rule["_id"])] = list(events)
            rule["events"] = self.events[str(rule["_id"])]
        collection = self.mongo.get_collection('events')
        tasks = []
        for rule in self.rules:
            tasks.append(
                asyncio.create_task(
                    get_rule_events(rule)
                )
            )
        await asyncio.gather(*tasks)

    async def get_formulas(self):
        formulas_collection = self.mongo.get_collection(coll_name="formulas")
        self.formulas = [row for row in await formulas_collection.find({}).to_list(length=None)]

    async def get_headers(self):
        collection = self.mongo.get_collection("headers")
        headers = await collection.find({"parent_id": ObjectId(self.flare_id)}).to_list(length=None)
        self.headers = list(headers)

    async def get_print_formulas(self):
        """get the formula with ids replaced with names"""
        for rule in self.rules:
            # should only be one match...
            formula = [
                formula for formula in self.formulas if rule["logic_id"] == formula["_id"]][0]
            regex = r"\[(.*?)\]"
            # list of ids (strings)
            placeholders = list(set(re.findall(regex, formula['formula'])))
            ph_mappings = {}
            regex = r"\(([^()]*)\)"
            for ph_id in placeholders:
                attr_found = list(set(re.findall(regex, ph_id)))
                if len(attr_found) > 0:
                    ph_id = ph_id.replace("(" + attr_found[0] + ")", "")
                ph_name = [formula["name"]
                           for formula in self.formulas if ph_id == str(formula["_id"])]
                if len(ph_name) == 0:
                    ph_name = [param["parameter"]
                               for param in self.parameters if ph_id == str(param["_id"])]
                    if len(ph_name) == 0:
                        ph_name = ["ERROR: could not find object from id"]
                        print("Error getting object from id in get_print_formulas")
                        print("Workaround used")
                ph_name = ph_name[0]
                ph_mappings[ph_id] = {"name": ph_name}
            for ph_id in ph_mappings:
                formula["formula"] = formula["formula"].replace(
                    ph_id, ph_mappings[ph_id]["name"])
                formula['formula'] = re.sub(r"\s\s+", " ", formula['formula'])
            rule["formula"] = formula

    async def get_target_value_name(self, rule):
        "Get the target value's name (for if values are requested)"
        oid = rule['value_id']
        mongo_filter = {"_id": ObjectId(oid)}
        collection = self.mongo.get_collection('formulas')
        response = await collection.find(mongo_filter).to_list(length=None)
        response = list(response)
        if (len(response) == 1):
            return response[0]["name"]
        collection = self.mongo.get_collection('parameters')
        response = await collection.find(mongo_filter).to_list(length=None)
        response = list(response)
        if (len(response) == 1):
            return response[0]["var_name"]
        return str(rule['value_id'])

    def get_s3_client(self):
        """ 
        Until we move over the buckets to the production environement... we need
        creds to access the dev enviro 
        """
        return boto3.client(
            's3',
            aws_access_key_id='AKIAUQAF22QZHT6RKRVG',
            aws_secret_access_key='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'
        )

    async def do_rule_calculations(self, rule):
        """min, max, and mean calculatons for event target value values"""
        to = rule["formula"]["to"]
        if to != "headers":
            for event in rule["events"]:
                event['value_min'] = min(event["values"])
                event['value_max'] = max(event["values"])
                event['value_mean'] = sum(
                    event["values"]) / len(event["values"])
            return

    async def group_header_values(self, rule):
        """
        for each rule
        set rule["events_by_header"]
        such that is it a list of lists
        where each list contains events grouped by header, header id, and 
        """
        # If its a to-headers formula for the rule
        rule["events_by_header"] = None
        if rule["formula"]["to"] != "headers":
            return

        groups = {}
        for event in rule["events"]:
            if event["parent_id"] not in groups:
                groups[event["parent_id"]] = []
            groups[event["parent_id"]].append(event)
        rule["events_by_header"] = [group for group in groups.values()]

    async def get_rule_per_header(self, rule):
        """
        Create a copy of the rule for each header
        Then replace events (for the flare) with the events for the specific header
        and TODO change the name to the header name??? anything else???
        """
        header_rules = []
        for index in range(len(self.headers)):
            header = self.headers[index]
            rule_copy = copy.deepcopy(rule)
            header_events = None
            # TODO, we can probably make accessing parent id cleaner here or coming up to here
            for group in rule_copy["events_by_header"]:
                parent_id = None
                try:
                    parent_id = group[0]["chunks"][0]["calculated"]["parent_id"]
                except:
                    print("Testing: No parent id found (yet not required)")
                if parent_id == header["_id"]:
                    header_events = group
                    break
            rule_copy["events"] = header_events
            rule_copy["header_name"] = header["name"].split(" ", 1)[1]
            header_rules.append(rule_copy)
        return header_rules

    async def write_to_excel(self, rule):
        to_header = False
        target = None

        if rule["formula"]["to"] == "headers":
            to_header = True
            title = f"{rule['header_name']} {rule['name']}"
            title = title[:30]
            target = self.wb[title]
        else:
            """ Get the correct rule tab """
            target = self.wb[rule['name']]

        start_date = parse(self.start_date)
        end_date = parse(self.end_date)
        total_days_in_period = (end_date - start_date).days + 1
        violation_days = 0
        target_date = start_date
        test_dates = []
        rule['events'] = [] if (rule['events'] == None) else rule['events']

        for _ in range(total_days_in_period):
            for event in rule['events']:
                if target_date.date() >= event['start'].date() and target_date.date() <= event['end'].date():
                    violation_days += 1
                    test_dates.append(target_date)
                    break
            target_date = target_date + timedelta(days=1)
        number_of_periods = sum([row['count'] for row in rule['events']])

        percent_violation_days = 0.0
        if violation_days > 0:
            percent_violation_days = violation_days/total_days_in_period

        # Fix merged Borders
        # https://stackoverflow.com/questions/38734044/losing-merged-cells-border-while-editing-excel-file-with-openpyxl
        for merged_cells in target.merged_cells.ranges:
            style = target.cell(merged_cells.min_row,
                                merged_cells.min_col)._style
            for col in range(merged_cells.min_col, merged_cells.max_col + 1):
                for row in range(merged_cells.min_row, merged_cells.max_row + 1):
                    target.cell(row, col)._style = style

        # print_json(rule)
        target.cell(column=6, row=2, value=datetime.now())
        target.cell(column=6, row=4, value=rule['name'])
        target.cell(column=6, row=5, value=rule["formula"]["name"])
        target.cell(column=6, row=6, value=rule['check_for'])
        target.cell(column=6, row=7, value=rule['with_values'])
        # TODO not sure if this is the intended functionality
        target.cell(column=6, row=8, value=await self.get_target_value_name(rule))

        event_name = ('Event: ' + rule['name']) if not to_header else (
            f'Event: {rule["name"]}, {rule["header_name"]}')
        unit_name = ('Unit: ' + self.event_unit) if not to_header else (
            f'Unit: {self.event_unit}, {rule["header_name"]}')
        target.cell(column=2, row=2, value='Event Summary: ' + rule['name'])
        target.cell(column=3, row=3, value=start_date)
        target.cell(column=3, row=4, value=end_date)
        target.cell(column=3, row=5, value=self.report_resolution)
        target.cell(column=3, row=6, value=len(rule['events']))
        target.cell(column=3, row=7, value=number_of_periods)
        target.cell(column=3, row=8, value=violation_days)
        target.cell(column=3, row=9, value=number_of_periods /
                    (total_days_in_period*96))
        target.cell(column=3, row=10, value=percent_violation_days)
        target.cell(column=2, row=12, value=event_name)
        target.cell(column=2, row=13, value=unit_name)
        target.cell(column=2, row=14,
                    value=f'{self.start_date} - {self.end_date}')
        target.cell(column=2, row=15, value='Formula: ' +
                    rule["formula"]["formula"])
        target.cell(column=2, row=16, value='Data Resolution: ' +
                    self.report_resolution)

        row_num = 18
        # print("num evetns: ", len(rule['events']))
        if (len(rule['events']) == 0):
            target.cell(column=2, row=row_num,
                        value="No deviations for this report")
        for row in range(len(rule['events'])):
            target.cell(column=2, row=row_num, value=row+1)
            target.cell(column=3, row=row_num, value=str(
                rule['events'][row]['start'])[:-3])
            target.cell(column=4, row=row_num, value=str(
                rule['events'][row]['end'])[:-3])
            target.cell(column=5, row=row_num,
                        value=rule['events'][row]['count'])

            # TODO This whole thing. calc values
            if 'value_min' in rule['events'][row]:
                target.cell(column=6, row=row_num,
                            value=rule['events'][row]['value_mean'])
                target.cell(column=7, row=row_num,
                            value=rule['events'][row]['value_max'])
                target.cell(column=8, row=row_num,
                            value=rule['events'][row]['value_min'])

                test_cell = target.cell(column=6, row=row_num)
                test_cell.number_format = '#,##0'
                test_cell = target.cell(column=7, row=row_num)
                test_cell.number_format = '#,##0'
                test_cell = target.cell(column=8, row=row_num)
                test_cell.number_format = '#,##0'
            row_num += 1

    # ====================================================================================================
    async def main(self):
        """
        - Get the flare that corresponds to the given flare id
        - a object of objects where they key corresponds to the rule and the value is a list of all the 'event' records that meet the criterias provided
        """
        await self.setup()

        """ Construct the output file name """
        generated = str(datetime.now().strftime("%Y-%m-%d_%H:%M:%S"))
        self.report_name = f'{self.flare["name"]}_Consolidated_Report_{self.start_date}-{self.end_date}__{generated}.xlsx'
        # print("report name:", self.report_name)

        """ Set the local and bucket paths """
        self.report_local_path = f"/tmp/{ self.report_name}"
        self.bucket_out_path = f"reports/{self.report_name}"
        self.bucket_get_path = f'templates/{self.template_name}'
        self.template_local_path = f"/tmp/{ self.template_name}"
        if DEV:
            self.template_local_path = f"{os.environ['dev_1_path']}{ self.template_name}"

        """ Set the event 'unit' (used in the final report) """
        self.event_unit = f'{self.flare["name"]} Flare ({self.flare["permit_id"]})'

        """ Download and load the template file """
        self.s3_client.download_file(
            self.bucket_name, self.bucket_get_path, self.template_local_path)
        wb = load_workbook(self.template_local_path)

        """ For each report determine the correct template and create a copy of the instance with the report name """
        self.report_names = []
        for report in self.rules:
            new_tab = None
            if report["formula"]["to"] == "headers":
                h_counter = 0
                for header in self.headers:
                    h_counter += 1
                    header_name = header["name"].split(" ", 1)[1]
                    if report['with_values']:
                        new_tab = wb.copy_worksheet(wb['with_values'])
                        title = f"{header_name} {report['name']}"
                        new_tab.title = title[:30]
                    else:
                        new_tab = wb.copy_worksheet(wb['without_values'])
                        title = f"{header_name} {report['name']}"
                        new_tab.title = title[:30]
                    self.report_names.append(new_tab.title)
            else:
                if report['with_values']:
                    new_tab = wb.copy_worksheet(wb['with_values'])
                    new_tab.title = report['name']
                else:
                    new_tab = wb.copy_worksheet(wb['without_values'])
                    new_tab.title = report['name']
                self.report_names.append(new_tab.title)

        """ Remove the original template tabs """
        for sheet in wb:
            if sheet.title not in self.report_names:
                wb.remove(wb[sheet.title])
        self.wb = wb
        # self.wb.save(filename = self.template_local_path)  #DELETE, FOR TESTING AND VIEWING RESULTS ONLY
        # return

        """ If "to headers" group events by header """
        tasks = []
        for rule in self.rules:
            rule["events_by_header"] = None
            tasks.append(
                asyncio.create_task(
                    self.group_header_values(rule)
                )
            )
        await asyncio.gather(*tasks)

        """Do calcs for rules w values"""
        tasks = []
        for rule in self.rules:
            # TODO uncomment this? maybe not.  save a touch of time. maybe remove and put within function
            if not rule["with_values"]:
                continue
            tasks.append(
                asyncio.create_task(
                    self.do_rule_calculations(rule)
                )
            )
        await asyncio.gather(*tasks)

        """ Write each report to their excel tab """
        for rule in self.rules:
            if rule["formula"]["to"] == "headers":
                header_rules = await self.get_rule_per_header(rule)
                for header_rule in header_rules:

                    await self.write_to_excel(header_rule)
            else:
                await self.write_to_excel(rule)

        """ Save the excel file to the local tmp dir """
        if DEV:
            # DEV PURPOSES ONLY
            self.wb.save(filename=self.template_local_path)
            return
        else:
            # WHAT WE ACTUAL WANT
            self.wb.save(filename=self.report_local_path)

        """ Upload to s3, and return the link """
        self.s3_client.upload_file(
            self.report_local_path, self.bucket_name, self.bucket_out_path)
        link = os.environ["bucket_link_path"] + self.bucket_out_path

        """Update job"""
        job_oid = ObjectId(self.mongo.event["job_id"])
        jobs_coll = self.mongo.get_collection("jobs")
        job = await jobs_coll.find_one({"_id": job_oid})
        start = job["start"]
        end = datetime.now()
        elapsed = end - start
        min, sec = divmod(elapsed.total_seconds(), 60)
        elapsed = f"{int(min)}:{int(sec)}"
        up = await jobs_coll.update_one({"_id": job_oid}, {
            "$set": {"end": datetime.now(), "progress": 100, "link": link, "elapsed" : elapsed},
        })
        print(f"for: {self.mongo.event['job_id']}; matched: {up.matched_count}; mod: {up.modified_count}")
        print(link)
        return link


async def main(event):
    widget = WidgetMain(event)
    link = await widget.main()
    return link


# event = {
#     "flare_id" : "5ef91f4b398f068e0fe3fadd",
#     "rule_ids" : ["5f36e717154d32cfa45cc87c", "5f3c13730c6ca810cf74a314", "5f3c138c0c6ca810cf74a315", "5f405a700e569f3178931068", "5f3d6e30e0813b05e8d1f6cc", "5f405a840e569f3178931069", "5f3f24340e569f3178931064", "5f3f24410e569f3178931065"], # , "5f3d6e30e0813b05e8d1f6cc"  <-the new header formula  #["5f36e717154d32cfa45cc87c"],
#     "start_date" : "2020-01-01",
#     "end_date" : "2020-03-31"
# }

# lambda_handler(event,None)

# FCC 5ef91f4b398f068e0fe3fadd
# LIU  5ef92163398f068e0fe3fade
# AG 5ef9218a398f068e0fe3fadf
