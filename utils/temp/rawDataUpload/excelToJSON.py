
# from flask import Flask, request, jsonify, session, abort, json
import json
from openpyxl import load_workbook
from datetime import datetime, timedelta, timezone
import formulas as sys_formulas
from itertools import chain
import itertools
import numpy as np
import re
import sys
import traceback
# app = Flask(__name__)
# app.secret_key = 'dljsaklqk24e21cjn!Ew@@dsa5'


def get_rows(sheet):
    for row in sheet.iter_rows():
        yield [cell.value for cell in row]


def getHeaderData(sheet):
    for row in sheet.iter_rows(max_row=6):
        yield [cell.value for cell in row]


def getSheetData(sheet):
    for row in sheet.iter_rows(min_row=6):
        yield [cell.value for cell in row]
    # yield row


workbook = None
dataRows = None
dataRows1 = None
headers = {}
cont = False


# @app.errorhandler(500)
# def resource_not_found(e):
#     return jsonify(error=str(e)), 500


# @app.route('/parseExcel', methods=['GET', 'POST'])
def parseExcel(data):
    global workbook
    global dataRows
    global dataRows1
    global headers
    global cont
    try:
        
        
        # x = {"hi": "erick"}
        # return jsonify(x)
        datapointCount = 0

        filename = data['fileName']
        outputFilePath = data['outputFile']

        date_format = '%Y-%m-%d %H:%M:%S'
        force = False
        if('force' in data and data['force'] == True):
            force = True

        if(workbook == None or force):
            workbook = load_workbook(filename=filename,
                                     data_only=True, read_only=True)
            sheet15 = workbook['15']
            dataRows = getSheetData(sheet15)
            tempHeaders = getHeaderData(sheet15)

            sheet1 = workbook['1']
            dataRows1 = getSheetData(sheet1)
            tempHeaders1 = getHeaderData(sheet1)

            rowNum = 1
            while(True):
                row = next(tempHeaders, None)
                if(row == None):
                    break
                if(rowNum == 1):
                    headers['piTagIds15'] = row #the pitagids... could be wrong from old db
                    rowNum = rowNum + 1
                    continue
                if(rowNum == 2):
                    headers['flareIds15'] = row #the flareids... could be wrong from old db
                    rowNum = rowNum + 1
                    continue
                if(rowNum == 5):
                    headers['PiTags15'] = row #the pitag names... should be right. no extension.
                    rowNum = rowNum + 1
                    continue
                rowNum = rowNum + 1

            rowNum = 1
            while(True):
                row = next(tempHeaders1, None)
                if(row == None):
                    break
                if(rowNum == 1):
                    headers['piTagIds1'] = row
                    rowNum = rowNum + 1
                    continue
                if(rowNum == 2):
                    headers['flareIds1'] = row
                    rowNum = rowNum + 1
                    continue
                if(rowNum == 5):
                    headers['PiTags1'] = row
                    rowNum = rowNum + 1
                    continue
                rowNum = rowNum + 1

        batchSize = data['batchSize']

        returnData = []

        rowData = []
        rowNum = 1
        while(True):
            if(rowNum == batchSize+1):
                break

            row = next(dataRows, None)
            if(row == None):
                break

            rowData.append(row)

            rowNum = rowNum + 1

        for column in range(len(headers['piTagIds15'])):
            print("15 column: ", column)
            if(column < 3):
                continue
            values = []
            dates = []
            piTag = headers['PiTags15'][column]
            for row in rowData:
                # dateTime = datetime.strptime(str(row[2]), date_format)  # .timestamp()
                dateTime = None
                try:
                    dateTime = datetime.strptime(str(row[2]), date_format)  # .timestamp()
                except Exception as e:
                    continue
                    x = {'a' : str(row[2]), 'b' : len(rowData)}
                    return jsonify(x)
                    return jsonify(str(row[2]))
                
                if LOWER_BOUND_DATE != None:
                    if (dateTime < LOWER_BOUND_DATE): 
                        # print(column, "15: ", dateTime)
                        continue
                    else: 
                        pass
                        # print(column, "15: ", dateTime, "INSERT")
                dateTime = dateTime.isoformat()
                values.append(row[column])
                #dates.append(dateTime - 18000)
                dates.append(dateTime)
            if(len(values) > 0):
                returnData.append({'p': piTag, 'v': values, 't': dates})
                datapointCount = datapointCount + len(values)

        rowData = []
        rowNum = 1
        while(True):
            if(rowNum == batchSize+1):
                break
            row = next(dataRows1, None)
            if(row == None):
                break

            rowData.append(row)

            rowNum = rowNum + 1

        for column in range(len(headers['piTagIds1'])):
            print("1 column: ", column)
            if(column < 3):
                continue
            values = []
            dates = []
            piTag = headers['PiTags1'][column]
            for row in rowData:
                dateTime = None
                try:
                    dateTime = datetime.strptime(str(row[2]), date_format)  # .timestamp()
                except Exception as e:
                    continue
                # dateTime = datetime.strptime(str(row[2]), date_format)
                # print(column, "1: ", dateTime)
                if LOWER_BOUND_DATE != None:
                    if (dateTime < LOWER_BOUND_DATE): 
                        # print(column, "15: ", dateTime)
                        continue
                    else: 
                        pass
                        # print(column, "15: ", dateTime, "INSERT")
                dateTime = dateTime.isoformat()
                values.append(row[column])
                #dates.append(dateTime - 18000)
                dates.append(dateTime)

            if(len(values) > 0):
                returnData.append({'p': piTag, 'v': values, 't': dates})
                datapointCount = datapointCount + len(values)

        if datapointCount > 0:
            print("writing file")
            with open(outputFilePath, 'w+') as outfile:
                json.dump(returnData, outfile)
            return "Complete"
            return jsonify({"status": 'complete'})
            # return jsonify({'item':len(rowData)})
            return jsonify(returnData)
        else:
            return "Complete"
        return jsonify({"status": 'complete'})

    except Exception as e:
        print("error: ", e)
        traceback.print_exc()
# 44580
LOWER_BOUND_DATE = datetime.fromtimestamp(1598832000000/1000.0)
x = parseExcel(
    {   
        'fileName':'C:\\Users\\evaquero\\Documents\\GitHub\\resuable-snippets\\files\\2020-8.xlsx', 
        'outputFile' : 'C:\\Users\\evaquero\\Documents\\GitHub\\resuable-snippets\\files\\output\\august_2020_end.json', 
        'batchSize' : 44646, 
        'force' : True
    })

print(x)