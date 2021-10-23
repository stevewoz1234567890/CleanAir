
from openpyxl import load_workbook
from datetime import datetime, timedelta, timezone
import formulas as sys_formulas
from itertools import chain
import itertools
import numpy as np
import re
import sys
import json


def get_rows(sheet):
    for row in sheet.iter_rows():
        yield [cell.value for cell in row]


def getHeaderData(sheet):
    for row in sheet.iter_rows(max_row=6):
        yield [cell.value for cell in row]


def getSheetData(sheet):
    for row in sheet.iter_rows(min_row=6):
        yield [cell.value for cell in row]
    # yield row


workbook = None
dataRows = None
dataRows1 = None
headers = {}


def parseExcel():
    global workbook
    global dataRows
    global dataRows1
    global headers

    # try:

    data = {
        "fileName": "2020Jan.xlsx",
        "batchSize": 1440,
        "force": False
    }

    filename = data['fileName']

    date_format = '%Y-%m-%d %H:%M:%S'
    force = False
    if('force' in data and data['force'] == True):
        force = True

    if(workbook == None or force):
        print('hit')
        workbook = load_workbook(filename=filename,
                                 data_only=True, read_only=True)
        #loaded = True
        sheet15 = workbook['15']
        dataRows = getSheetData(sheet15)
        tempHeaders = getHeaderData(sheet15)
        print(tempHeaders)

        sheet1 = workbook['1']
        dataRows1 = getSheetData(sheet1)
        tempHeaders1 = getHeaderData(sheet1)

        rowNum = 1
        while(True):
            row = next(tempHeaders, None)
            if(row == None):
                break
            if(rowNum == 1):
                headers['piTagIds15'] = row
                rowNum = rowNum + 1
                continue
            if(rowNum == 2):
                headers['flareIds15'] = row
                rowNum = rowNum + 1
                continue
            if(rowNum == 5):
                headers['PiTags15'] = row
                rowNum = rowNum + 1
                continue
            rowNum = rowNum + 1
        rowNum = 1
        while(True):
            row = next(tempHeaders1, None)
            if(row == None):
                break
            if(rowNum == 1):
                headers['piTagIds1'] = row
                rowNum = rowNum + 1
                continue
            if(rowNum == 2):
                headers['flareIds1'] = row
                rowNum = rowNum + 1
                continue
            if(rowNum == 5):
                headers['PiTags1'] = row
                rowNum = rowNum + 1
                continue
            rowNum = rowNum + 1
    batchSize = data['batchSize']
    rowNum = 1
    returnData = []

    rowData = []
    while(True):
        row = next(dataRows, None)
        if(row == None):
            break
        if(rowNum == batchSize+1):
            break
        rowData.append(row)
        print(json.dumps(row, indent=4, default=str))
        rowNum = rowNum + 1

    for column in range(len(headers['piTagIds15'])):
        if(column < 3):
            continue
        values = []
        dates = []
        piTag = headers['PiTags15'][column]
        for row in rowData:
            dateTime = datetime.strptime(
                str(row[2]), date_format).timestamp()
            values.append(row[column])
            dates.append(dateTime - 18000)
        returnData.append({'p': piTag, 'v': values, 't': dates})

    rowData = []
    rowNum = 1
    while(True):
        row = next(dataRows1, None)
        if(row == None):
            break
        if(rowNum == batchSize+1):
            break
        rowData.append(row)
        rowNum = rowNum + 1

    for column in range(len(headers['piTagIds1'])):
        if(column < 3):
            continue
        values = []
        dates = []
        piTag = headers['PiTags1'][column]
        for row in rowData:
            dateTime = datetime.strptime(
                str(row[2]), date_format).timestamp()
            values.append(row[column])
            dates.append(dateTime - 18000)
        returnData.append({'p': piTag, 'v': values, 't': dates})
    print(returnData)

    # except Exception as e:
    #     print(e)


if __name__ == "__main__":
    parseExcel()
