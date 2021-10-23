
from flask import Flask, request, jsonify, session, abort
from openpyxl import load_workbook
from datetime import datetime, timedelta, timezone
import formulas as sys_formulas
from itertools import chain
import itertools
import numpy as np
import re
import sys
app = Flask(__name__)
app.secret_key = 'dljsaklqk24e21cjn!Ew@@dsa5'


def get_rows(sheet):
    for row in sheet.iter_rows():
        yield [cell.value for cell in row]


def getHeaderData(sheet):
    for row in sheet.iter_rows(max_row=6):
        yield [cell.value for cell in row]


def getSheetData(sheet):
    for row in sheet.iter_rows(min_row=6):
        yield [cell.value for cell in row]
    # yield row


workbook = None
dataRows = None
dataRows1 = None
headers = {}
cont = False


@app.errorhandler(500)
def resource_not_found(e):
    return jsonify(error=str(e)), 500


@app.route('/parseExcel', methods=['GET', 'POST'])
def parseExcel():
    global workbook
    global dataRows
    global dataRows1
    global headers
    global cont
    try:
        datapointCount = 0
        data = request.json

        filename = 'utils/python/' + data['fileName']

        date_format = '%Y-%m-%d %H:%M:%S'
        force = False
        if('force' in data and data['force'] == True):
            force = True

        if(workbook == None or force):
            workbook = load_workbook(filename=filename,
                                     data_only=True, read_only=True)
            sheet15 = workbook['15']
            dataRows = getSheetData(sheet15)
            tempHeaders = getHeaderData(sheet15)

            sheet1 = workbook['1']
            dataRows1 = getSheetData(sheet1)
            tempHeaders1 = getHeaderData(sheet1)

            rowNum = 1
            while(True):
                row = next(tempHeaders, None)
                if(row == None):
                    break
                if(rowNum == 1):
                    headers['piTagIds15'] = row
                    rowNum = rowNum + 1
                    continue
                if(rowNum == 2):
                    headers['flareIds15'] = row
                    rowNum = rowNum + 1
                    continue
                if(rowNum == 5):
                    headers['PiTags15'] = row
                    rowNum = rowNum + 1
                    continue
                rowNum = rowNum + 1

            rowNum = 1
            while(True):
                row = next(tempHeaders1, None)
                if(row == None):
                    break
                if(rowNum == 1):
                    headers['piTagIds1'] = row
                    rowNum = rowNum + 1
                    continue
                if(rowNum == 2):
                    headers['flareIds1'] = row
                    rowNum = rowNum + 1
                    continue
                if(rowNum == 5):
                    headers['PiTags1'] = row
                    rowNum = rowNum + 1
                    continue
                rowNum = rowNum + 1

        batchSize = data['batchSize']

        returnData = []

        rowData = []
        rowNum = 1
        while(True):
            if(rowNum == batchSize+1):
                break

            row = next(dataRows, None)
            if(row == None):
                break

            rowData.append(row)

            rowNum = rowNum + 1

        for column in range(len(headers['piTagIds15'])):
            if(column < 3):
                continue
            values = []
            dates = []
            piTag = headers['PiTags15'][column]
            for row in rowData:
                dateTime = datetime.strptime(
                    str(row[2]), date_format)  # .timestamp()
                values.append(row[column])
                #dates.append(dateTime - 18000)
                dates.append(dateTime)
            if(len(values) > 0):
                returnData.append({'p': piTag, 'v': values, 't': dates})
                datapointCount = datapointCount + len(values)

        rowData = []
        rowNum = 1
        while(True):
            if(rowNum == batchSize+1):
                break
            row = next(dataRows1, None)
            if(row == None):
                break

            rowData.append(row)

            rowNum = rowNum + 1

        for column in range(len(headers['piTagIds1'])):
            if(column < 3):
                continue
            values = []
            dates = []
            piTag = headers['PiTags1'][column]
            for row in rowData:
                dateTime = datetime.strptime(
                    str(row[2]), date_format)  # .timestamp()
                values.append(row[column])
                #dates.append(dateTime - 18000)
                dates.append(dateTime)

            if(len(values) > 0):
                returnData.append({'p': piTag, 'v': values, 't': dates})
                datapointCount = datapointCount + len(values)

        cont = True
        if datapointCount > 0:
            return jsonify(returnData)
        else:
            return jsonify({"status": 'complete'})

    except Exception as e:
        print(e)
        return jsonify({'error': e})


@app.route('/hello', methods=['GET', 'POST'])
def hello():
    data = request.json
    return jsonify({"testing": 'test3'})


@app.route('/formula', methods=['POST'])
def getFormulaValue():
    try:
        data = request.json
        inFormula = cleanFormula(data['formula'])
        formula = sys_formulas.Parser().ast(inFormula)[1].compile()
        inputs = list(formula.inputs)
        variables = cleanVariables(data['variables'])
        if "NULL" in inputs:
            variables['NULL'] = None

        response = formula(**variables)
        cleanResp = parse_return_value(response)
        return jsonify({"value": cleanResp, "rawRes": str(response), "variables": variables})
    except Exception as e:
        # print(e)
        # return jsonify({'error': e})
        abort(500, description=e)
        # return jsonify(e.to_dict())


def cleanVariables(variables):
    newVars = {}
    for variable in variables:
        newVar = variable.replace(
            "[", '').replace("]", '').replace(" ", "_")
        newVars[newVar.upper()] = variables[variable]
    return newVars


def addDefaultVars(variables):
    pass
    variables['NULL'] = None


def cleanFormula(formula):
    regex = r"\[(.*?)\]"
    variables = list(set(re.findall(regex, formula)))
    formula = re.sub("\s\s+", " ", formula)
    formula = formula.lstrip()
    if formula[0] != "=":
        formula = "=" + formula

    for variable in variables:
        clean_var = variable.replace(
            "[", '').replace("]", '').replace(" ", "_")
        formula = formula.replace("[" + variable + "]", clean_var.upper())
    return formula


def parse_return_value(value=None):
    if "#VALUE!" in str(value):
        return None
    if "#N/A" in str(value):
        return None
    if isinstance(value, float):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, list):
        pass
        #print("LIST HIT")
    if type(value) == np.bool_:
        value = value.tolist()
        return value
    if type(value) == np.ndarray:
        value = value.tolist()
        new_list = list(itertools.chain(*value))
        # _json(new_list)
        return new_list
    if type(value) == sys_formulas.functions.Array:
        value = value.tolist()

        if isinstance(value, float):
            return value
        if isinstance(value, int):
            return value
        if isinstance(value, bool):
            return value
        if isinstance(value, list):
            # print("HIT")
            if len(value) == 1:
                new_list = list(itertools.chain(*value))[0]
                if isinstance(new_list, float):
                    return new_list
                if isinstance(new_list, int):
                    return new_list
                if isinstance(new_list, bool):
                    return new_list
                return new_list[0]

            return value

            # print(value)


if __name__ == "__main__":
    app.run(port=sys.argv[1], debug=True)
