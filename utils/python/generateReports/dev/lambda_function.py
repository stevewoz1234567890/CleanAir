import json #standard
import sys #standard
import asyncio #standard
from datetime import datetime, timedelta, timezone #standard
from statistics import mean #standard
import re #standard
from collections import defaultdict  #standard # seperate data by header
import copy #standard
import traceback #standard
from bson import ObjectId
import pytz
from openpyxl import load_workbook, Workbook
import boto3
import motor.motor_asyncio
from dateutil.parser import parse, parserinfo
import functools #for reduce

parserInfoInput = parserinfo(False, True)

if __name__ == '__main__':
    import config as os
else:
    import os

DEV = True if os.environ["env"] == "local_dev" else False # specific for erick's env. Turn off with False


"""
TODO
- Add feature for events that extend beyond start and end times but fall within that timespan

- calculate values for 'to header' formulas

"""

def utc_to_local(utc_dt, local_tz):
    local_dt = utc_dt.replace(tzinfo=pytz.utc).astimezone(local_tz)
    return local_tz.normalize(local_dt) # helps with dst changes

def print_json(data):
    print(json.dumps(data, indent=4, default=str))


def _json(data):
    print(json.dumps(data, indent=4, default=str))


def lambda_handler(event=None, context=None):
    print("event: ", event)
    message = "successfully created report"
    response = {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Headers': 'Content-Type,X-Api-Key',
            'Access-Control-Allow-Origin': '*',
        }
    }
    try:
        #try starts
        link = asyncio.run(
            main(event)
        )
        #try ends
    except Exception as err:
        #except starts
        print("ERROR: ", err)
        traceback.print_exc()
        response["statusCode"] = 500
        link = None
        message = "Error occurred. Talk to Dev Team"
        #except ends
    
    response["body"] = json.dumps(
        {"link": link, "message": message}, default=str)
    print(response)
    return response


class Mongo():
    def __init__(self, event, database_name=os.environ["db_name"]):
        self.database_name = database_name
        self.event = event
        self.conn_string = os.environ["conn_string"]
        self.client = motor.motor_asyncio.AsyncIOMotorClient(self.conn_string)

    def get_collection(self, coll_name):
        db = self.client[self.database_name]
        coll = db[coll_name]
        return coll


class ReportGenerator():
    def __init__(self, event):
        self.mongo = Mongo(event)
        self.local_timezone = event['timezone']
        self.pytz_local_tz = pytz.timezone(self.local_timezone)
        self.debug = event['debug']
        self.flare_id = event['flare_id']  # the flare id string
        self.flare = None  # Flare object
        self.rule_pairs = event['rule_ids']
        self.rule_ids = [pair['_id'] for pair in event['rule_ids']]  # the id of each report we want
        self.rules = None  # Rule objects
        self.events = {}
        self.s3_client = self.get_s3_client()
        self.bucket_name = 'flare-reporting'
        self.start_date = event['start_date']  # start date
        self.end_date = event['end_date']  # end date
        self.template_name = 'report_template.xlsx'
        self.headers = []
        self.job_id = event['job_id']

    async def setup(self):
        self.rules_type_map = {}
        for pair in self.rule_pairs:
            self.rules_type_map[pair['_id']] = pair['type']

        await self.get_times()
        await self.get_parameters()
        await self.get_flare()
        await self.get_rules()
        await self.get_events()
        await self.get_formulas()
        await self.get_headers()
        await self.assign_formula_to_rule()
        # await self.get_print_formulas()

    async def get_times(self):
        local = self.pytz_local_tz
        naive_dt = parse(self.start_date, parserInfoInput)
        self.local_start_dt = local.localize(naive_dt, is_dst=None)
        self.utc_start_dt = self.local_start_dt.astimezone(pytz.utc)

        naive_dt = parse(self.end_date, parserInfoInput)
        self.local_end_dt = local.localize(naive_dt, is_dst=None) + timedelta(days=1)
        self.utc_end_dt = self.local_end_dt.astimezone(pytz.utc)


    async def get_parameters(self):
        collection = self.mongo.get_collection(coll_name="parameters")
        self.parameters = [row for row in await collection.find({}).to_list(length=None)]

    async def get_flare(self):
        flares_collection = self.mongo.get_collection(coll_name="flares")
        self.flare = await flares_collection.find_one({"_id": ObjectId(self.flare_id)})

    async def get_rules(self):
        rules_collection = self.mongo.get_collection(coll_name="eventrules")
        num_rules_collection = self.mongo.get_collection(coll_name="numericEventRules")
        rules = []
        for id in self.rule_ids:
            if id == "VISIBLE_EMISSIONS":
                rule = {
                    "_id" : "VISIBLE_EMISSIONS",
                    "checkForValue" : None,
                    "name" : "Visible Emissions",
                    "resolution" : 1,
                    "chunkSize" : 0,
                    "sensitivity" : 0,
                    "formula" : "VISIBLE_EMISSIONS",
                    "checkFor" : True,
                    "withValues" : False,
                    "type" : "special"
                }
                rules.append(rule)
            else:
                if (self.rules_type_map[id] == "numeric"):
                    rule = await num_rules_collection.find_one({"_id": ObjectId(id)})
                    if rule == None:
                        print("No matching rule was found")
                    else:
                        rule['type'] = "numeric"
                        rule["checkForValue"] = None
                        rule["chunkSize"] = 0
                        rule["sensitivity"] = 0
                        rule["checkFor"] = True
                        rule["withValues"] = False
                        rules.append(rule)
                elif (self.rules_type_map[id] == "boolean"): 
                    rule = await rules_collection.find_one({"_id": ObjectId(id)})
                    if rule == None:
                        print("No matching rule was found")
                    else:
                        rule['type'] = "boolean"
                        rules.append(rule)
        self.rules = rules

    def dailyDocsToEvents(self, rule, docs):
        # This has limited functionality, it only work for units of 1 day
        if (rule['actionPeriod'] != "day" and rule['actionPeriodLength'] != 1 and rule['actionPeriodAction'] != 'rolling'):
            raise ValueError('Reports for numeric event rules beyond 1-day-rolling-sum is not supported')
        def docToEvent(doc):
            newDoc = {}
            newDoc['isEvent'] = False
            values = [(0 if datum['value'] == None else datum['value']) for datum in doc['data']]
            sum = functools.reduce(lambda a, b: a+b, values)
            if (rule['actionOperation'] == ">"):
                if (sum > rule['actionValue']): newDoc['isEvent'] = True
            elif (rule['actionOperation'] == "<"):
                if (sum > rule['actionValue']): newDoc['isEvent'] = True
            elif (rule['actionOperation'] == ">="):
                if (sum > rule['actionValue']): newDoc['isEvent'] = True
            elif (rule['actionOperation'] == "<="):
                if (sum > rule['actionValue']): newDoc['isEvent'] = True
            elif (rule['actionOperation'] == "="): #I don't think '=' should be an option...
                if (sum > rule['actionValue']): newDoc['isEvent'] = True
            if (newDoc['isEvent']):
                newDoc['start'] = utc_to_local(doc['start'], self.pytz_local_tz)
                newDoc['end'] = utc_to_local(doc['end'], self.pytz_local_tz)
                newDoc['flare'] = doc['flare']
                newDoc['header'] = doc['header']
                # newDoc['flare'] = 


            return newDoc
        events = map(docToEvent, docs)
        events = filter(lambda event: event['isEvent'], events)
        return events

    async def get_events(self):
        """Get the events for each rule"""
        async def get_rule_events(rule): #TODO this would be more efficient if it was one call rather than one per rule
            """Get the events for a single rule"""
            if rule["type"] == "special":
                if rule["_id"] == "VISIBLE_EMISSIONS":
                    mongo_filter = {
                        'flare' : ObjectId(self.flare_id),
                        'endDate': { #so that we can capture events that start before our limit, but continue into our selected daterange
                            "$gt": self.utc_start_dt
                        },
                        'startDate': {
                            "$lt": self.utc_end_dt
                        }
                    }
                    emissions_coll = self.mongo.get_collection("visibleemissions")
                    events =  await emissions_coll.find(mongo_filter).to_list(length=None)
                    events = list(events)
                    for event in events:
                        event['start'] = utc_to_local(event['startDate'], self.pytz_local_tz)
                        event['end'] = utc_to_local(event['endDate'], self.pytz_local_tz)
                        time_delta = (event['endDate'] - event['startDate'])
                        total_seconds = time_delta.total_seconds()
                        minutes = total_seconds/60
                        if minutes == 0:
                            event['chunks'] = []
                        else:
                            event['chunks'] = [True for i in range(int(minutes))]
                    self.events[str(rule["_id"])] = events
                    rule["events"] = self.events[str(rule["_id"])]
            elif rule["type"] == "boolean":
                coll_name = "debugEvents" if self.debug else "events"
                collection = self.mongo.get_collection(coll_name)
                mongo_filter = {
                    'flare': ObjectId(self.flare_id),
                    'eventRule': rule["_id"],
                    'end': { #so that we can capture events that start before our limit, but continue into our selected daterange
                        "$gt": self.utc_start_dt
                    },
                    'start': {
                        "$lt": self.utc_end_dt
                    }
                }
                events = await collection.find(mongo_filter).to_list(length=None)
                events = list(events)
                for event in events:
                    event['start'] = utc_to_local(event['start'], self.pytz_local_tz)
                    event['end'] = utc_to_local(event['end'], self.pytz_local_tz)
                self.events[str(rule["_id"])] = events
                rule["events"] = self.events[str(rule["_id"])]
            elif rule["type"] == "numeric":
                print("NUMERIC RULE")
                collection = self.mongo.get_collection("dailyNumericData")
                mongo_filter = {
                    'parameter' : rule["parameter"],
                    'flare' : ObjectId(self.flare_id),
                    'start' : {
                        '$gte' : self.utc_start_dt,
                        '$lte' : self.utc_end_dt
                    }
                }
                # print(self.utc_start_dt)
                # print(self.utc_end_dt)
                # exit()
                events = await collection.find(mongo_filter).to_list(length=None)
                events = list(events)
                events = list(self.dailyDocsToEvents(rule, events))
                print("PROCESSED EVENTS: ", events)
                # exit()
                self.events[str(rule["_id"])] = events
                rule["events"] = self.events[str(rule["_id"])]
            else:
                print("rule does not have a valid type")
                print("rule: ", rule)
        
        tasks = []
        for rule in self.rules:
            tasks.append(
                asyncio.create_task(
                    get_rule_events(rule)
                )
            )
        await asyncio.gather(*tasks)

    async def get_formulas(self):
        formulas_collection = self.mongo.get_collection(coll_name="formulas")
        self.formulas = [row for row in await formulas_collection.find({}).to_list(length=None)]

    async def get_headers(self):
        collection = self.mongo.get_collection("headers")
        headers = await collection.find({"flare": ObjectId(self.flare_id)}).to_list(length=None)
        self.headers = list(headers)

    async def assign_formula_to_rule(self):
        for rule in self.rules:
            if (rule['type'] == "special"):
                if (rule["_id"] == "VISIBLE_EMISSIONS"):
                    rule["formula"] = {
                        "name" : "Visible Emissions",
                        "formula" : "Not Applicable",
                        "to" : "flare"
                    }
                    continue
            elif (rule['type'] == "boolean"):
                formula_id = rule['formula']
                formula = [formula for formula in self.formulas if formula_id == formula["_id"]][0]
                rule['formula'] = formula
            elif (rule['type'] == "numeric"):
                parameter_id = rule['parameter']
                if (rule['parameterType'] == "formula"):
                    formula = [formula for formula in self.formulas if parameter_id == formula["_id"]][0]
                    print(formula)
                    rule['formula'] = formula
            if "-" in rule["name"]:
                    # This was a formatting fix for the rules with labels (of category, for example) at front of name. 
                    # It only should effect ouput apperance not functionality
                    name_parts = rule["name"].split("-",1)
                    rule['name'] = name_parts[1]

    async def get_print_formulas(self):
        """get the formula with ids replaced with names"""
        for rule in self.rules:
            # should only be one match...
            formula = [
                formula for formula in self.formulas if rule["logic_id"] == formula["_id"]][0]
            regex = r"\[(.*?)\]"
            # list of ids (strings)
            placeholders = list(set(re.findall(regex, formula['formula'])))
            ph_mappings = {}
            regex = r"\(([^()]*)\)"
            for ph_id in placeholders:
                attr_found = list(set(re.findall(regex, ph_id)))
                if len(attr_found) > 0:
                    ph_id = ph_id.replace("(" + attr_found[0] + ")", "")
                ph_name = [formula["name"]
                           for formula in self.formulas if ph_id == str(formula["_id"])]
                if len(ph_name) == 0:
                    ph_name = [param["parameter"]
                               for param in self.parameters if ph_id == str(param["_id"])]
                    if len(ph_name) == 0:
                        ph_name = ["ERROR: could not find object from id"]
                        print("Error getting object from id in get_print_formulas")
                        print("Workaround used")
                ph_name = ph_name[0]
                ph_mappings[ph_id] = {"name": ph_name}
            for ph_id in ph_mappings:
                formula["formula"] = formula["formula"].replace(
                    ph_id, ph_mappings[ph_id]["name"])
                formula['formula'] = re.sub(r"\s\s+", " ", formula['formula'])
            rule["formula"] = formula

    async def get_target_value_name(self, rule):
        "Get the target value's name (for if values are requested)"
        oid = rule['checkForValue']
        mongo_filter = {"_id": oid}
        collection = self.mongo.get_collection('formulas')
        response = await collection.find(mongo_filter).to_list(length=None)
        response = list(response)
        if (len(response) == 1):
            return response[0]["name"]
        collection = self.mongo.get_collection('parameters')
        response = await collection.find(mongo_filter).to_list(length=None)
        response = list(response)
        if (len(response) == 1):
            return response[0]["name"]
        return str(rule['checkForValue'])

    def get_s3_client(self):
        """ 
        Until we move over the buckets to the production environement... we need
        creds to access the dev enviro 
        """
        return boto3.client(
            's3',
            aws_access_key_id='AKIAUQAF22QZHT6RKRVG',
            aws_secret_access_key='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'
        )

    async def do_rule_calculations(self, rule):
        """min, max, and mean calculatons for event target value values"""
        for event in rule["events"]:
            values = [entry['value'] for entry in event['values'] if entry['value'] != None and type(entry['value']) != type("")]
            noValidData = False
            if (len(values) == 0):
                noValidData = True
            event['value_min'] = "N/A" if noValidData else min(values)
            event['value_max'] = "N/A" if noValidData else max(values)
            event['value_mean'] = "N/A" if noValidData else (sum(values) / len(values))
        return

    async def group_header_values(self, rule):
        """
        for each rule
        set rule["events_by_header"]
        such that is it a list of lists
        where each list contains events grouped by header, header id, and 
        """
        # If its a to-headers formula for the rule
        rule["events_by_header"] = None
        if rule["formula"]["to"] != "headers":
            return

        groups = {}
        for event in rule["events"]:
            if event["header"] not in groups: #TODO: it seems like it would be better to init all the header arrays at once and if there is no event for it, then use the empty array rather than checking if the key exists at all
                groups[event["header"]] = []
            groups[event["header"]].append(event)
        rule["events_by_header"] = [group for group in groups.values()]

    async def get_rule_per_header(self, rule):
        """
        Create a copy of the rule for each header
        Then replace events (for the flare) with the events for the specific header
        """
        #TODO, we can probably skip doing a copy for single-header flares
        header_rules = []
        for index in range(len(self.headers)):
            header = self.headers[index]
            rule_copy = copy.deepcopy(rule)
            header_events = None
            # TODO, we can probably make accessing parent id cleaner here or coming up to here
            for group in rule_copy["events_by_header"]:
                parent_id = None
                try:
                    parent_id = group[0]["header"]
                except:
                    print("Testing: No parent id found (yet not required)")
                if parent_id == header["_id"]:
                    header_events = group
                    break
            rule_copy["events"] = header_events
            rule_copy["header_name"] = header["name"].split(" ", 1)[1]
            header_rules.append(rule_copy)
        return header_rules

    async def write_to_excel(self, rule):
        to_header = False
        target = None
        resolution = str(rule['resolution']) + " minutes"
        total_periods_per_day = 96 if rule['resolution'] == 15 else 1440
        total_periods_per_day = 1440 if rule['resolution'] == 1 else 1
        total_periods_per_day = 1 if rule['type'] == 'numeric' else 1

        if rule["formula"]["to"] == "headers":
            to_header = True
            title = f"{rule['header_name']} {rule['name']}"
            title = title[:30]
            target = self.wb[title]
        else:
            """ Get the correct rule tab """
            title = f"{self.flare['name']} {rule['name']}"
            title = title[:30]
            target = self.wb[title]

        start_date = self.local_start_dt #parse(self.start_date, parserInfoInput) #TODO we don't need to be parsing for each reporting tab
        end_date = self.local_end_dt - timedelta(days=1)#parse(self.end_date, parserInfoInput)
        total_days_in_period = (end_date - start_date).days + 1
        violation_days = 0
        target_date = start_date
        test_dates = []
        rule['events'] = [] if (rule['events'] == None) else rule['events']
        rule['events'].reverse()     

        for _ in range(total_days_in_period): #TODO, i'm not sure I follow this. Seems like it would always get 100% of the days?
            for event in rule['events']:
                # print("CHECK S: ", event['start'], type(event['start']))
                # print("CHECK E: ", event['end'], type(event['end']))
                if target_date.date() >= event['start'].date() and target_date.date() <= event['end'].date(): #TODO why using .date()
                    violation_days += 1
                    test_dates.append(target_date)
                    break
            target_date = target_date + timedelta(days=1)
        
        num_events = len(rule['events'])
        number_of_periods = num_events if rule['type'] == 'numeric' else sum([len(row['chunks']) for row in rule['events']])

        percent_violation_days = 0.0
        if violation_days > 0:
            percent_violation_days = violation_days/total_days_in_period

        # Fix merged Borders
        # https://stackoverflow.com/questions/38734044/losing-merged-cells-border-while-editing-excel-file-with-openpyxl
        for merged_cells in target.merged_cells.ranges:
            style = target.cell(merged_cells.min_row,
                                merged_cells.min_col)._style
            for col in range(merged_cells.min_col, merged_cells.max_col + 1):
                for row in range(merged_cells.min_row, merged_cells.max_row + 1):
                    target.cell(row, col)._style = style

        # print_json(rule)
        target.cell(column=6, row=2, value=datetime.now())
        target.cell(column=6, row=4, value=rule['name'])
        target.cell(column=6, row=5, value=rule["formula"]["name"])
        target.cell(column=6, row=6, value=rule['checkFor'])
        target.cell(column=6, row=7, value=rule['withValues'])
        # TODO not sure if this is the intended functionality
        target.cell(column=6, row=8, value=await self.get_target_value_name(rule))

        event_name = ('Event: ' + rule['name']) if not to_header else (
            f'Event: {rule["name"]}, {rule["header_name"]}')
        unit_name = ('Unit: ' + self.event_unit) if not to_header else (
            f'Unit: {self.event_unit}, {rule["header_name"]}')
        target.cell(column=2, row=2, value='Event Summary: ' + rule['name'])
        target.cell(column=3, row=3, value=start_date)
        target.cell(column=3, row=4, value=end_date)
        target.cell(column=3, row=5, value=resolution)
        target.cell(column=3, row=6, value=len(rule['events']))
        target.cell(column=3, row=7, value=number_of_periods)
        target.cell(column=3, row=8, value=violation_days)
        target.cell(column=3, row=9, value=number_of_periods /
                    (total_days_in_period*total_periods_per_day))
        target.cell(column=3, row=10, value=percent_violation_days)
        target.cell(column=2, row=12, value=event_name)
        target.cell(column=2, row=13, value=unit_name)
        target.cell(column=2, row=14,
                    value=f'{self.start_date} - {self.end_date}') #TODO, I think this will need to be fixed because I doent see permentant conversions
        num_rule_display = f'{rule["actionPeriodLength"]} {rule["actionPeriod"]} {rule["actionPeriodAction"]} {rule["actionOperation"]} {rule["actionInequality"]} {rule["actionValue"]}' if rule['type'] == 'numeric' else None
        formula_display = f'Condition: {num_rule_display}' if rule['type'] == 'numeric' else f'Formula: {rule["formula"]["formula"]}'
        target.cell(column=2, row=15, value=formula_display)
        target.cell(column=2, row=16, value='Data Resolution: ' +
                    resolution)

        row_num = 18
        # print("num evetns: ", len(rule['events']))
        # rule['events'].reverse()
        if (len(rule['events']) == 0):
            target.cell(column=2, row=row_num,
                        value="No deviations for this report")
        for row in range(len(rule['events'])):
            target.cell(column=2, row=row_num, value=row+1)
            target.cell(column=3, row=row_num, value=str(
                rule['events'][row]['start'])[:-6])
            target.cell(column=4, row=row_num, value=str(
                rule['events'][row]['end'])[:-6])
            event_duration = 1 if rule['type'] == 'numeric' else len(rule['events'][row]['chunks'])
            target.cell(column=5, row=row_num,
                        value=event_duration)

            if 'value_min' in rule['events'][row]:
                target.cell(column=6, row=row_num,
                            value=rule['events'][row]['value_mean'])
                target.cell(column=7, row=row_num,
                            value=rule['events'][row]['value_max'])
                target.cell(column=8, row=row_num,
                            value=rule['events'][row]['value_min'])

                #TODO, cant recall what the below does exactly, although its obvious its some kind of formatting
                test_cell = target.cell(column=6, row=row_num)
                test_cell.number_format = '#,##0'
                test_cell = target.cell(column=7, row=row_num)
                test_cell.number_format = '#,##0'
                test_cell = target.cell(column=8, row=row_num)
                test_cell.number_format = '#,##0'
            row_num += 1

    async def main(self):
        """
        - Get the flare that corresponds to the given flare id
        - a object of objects where they key corresponds to the rule and the value is a list of all the 'event' records that meet the criterias provided
        """
        # print("rule ids before: ", self.rule_ids)
        # self.rule_ids = list(filter(lambda x: x != "VISIBLE_EMISSIONS", self.rule_ids))
        # print("rule ids after: ", self.rule_ids)
        await self.setup()

        """ Construct the output file name """
        generated = str(datetime.now().strftime("%Y-%m-%dT%H-%M-%S"))
        self.report_name = f'{self.flare["name"]}_Consolidated_Report_{self.start_date.replace("/","-")}-{self.end_date.replace("/","-")}__{generated}.xlsx'
        # self.report_name = f'{self.flare["name"]}_Consolidated_Report.xlsx'
        # print("report name:", self.report_name)

        """ Set the local and bucket paths """
        self.report_local_path = f"/tmp/{self.report_name}"
        self.bucket_out_path = f"reports/{self.report_name}"
        self.bucket_get_path = f'templates/{self.template_name}'
        self.template_local_path = f"/tmp/{ self.template_name}"
        if DEV:
            self.template_local_path = f"{os.environ['dev_1_path']}{self.template_name}"
            self.local_save = f"{os.environ['dev_1_path']}OUTPUT_{ self.report_name}"

        """ Set the events' (flare) 'unit' displayed in the output report """
        self.event_unit = f'{self.flare["name"]} Flare ({self.flare["permitId"]})'

        """ Download and load the template file """
        if not DEV:
            self.s3_client.download_file(
                self.bucket_name, self.bucket_get_path, self.template_local_path)
        wb = load_workbook(self.template_local_path)

        """ For each report determine the correct template and create a copy of the instance with the report name """
        self.report_names = []

        for report in self.rules:
            new_tab = None
            if report["formula"]["to"] == "headers":
                h_counter = 0
                for header in self.headers:
                    h_counter += 1
                    header_name = header["name"].split(" ", 1)[1]
                    title = f"{header_name} {report['name']}"
                    if report['withValues']:
                        new_tab = wb.copy_worksheet(wb['with_values'])
                        new_tab.title = title[:30] #tab titles have a character limit
                    else:
                        new_tab = wb.copy_worksheet(wb['without_values'])
                        new_tab.title = title[:30]
                    self.report_names.append(new_tab.title)
            else:
                title = f"{self.flare['name']} {report['name']}"
                if report['withValues']:
                    new_tab = wb.copy_worksheet(wb['with_values'])
                    new_tab.title = title[:30]
                else:
                    new_tab = wb.copy_worksheet(wb['without_values'])
                    new_tab.title = title[:30]
                self.report_names.append(new_tab.title)
        
        

        """ Remove the original template tabs """
        for sheet in wb:
            if sheet.title not in self.report_names:
                wb.remove(wb[sheet.title])
        self.wb = wb
        # self.wb.save(filename = self.template_local_path)  #DELETE, FOR TESTING AND VIEWING RESULTS ONLY
        # return

        """ If "to headers" group events by header """
        tasks = []
        for rule in self.rules:
            rule["events_by_header"] = None
            tasks.append(
                asyncio.create_task(
                    self.group_header_values(rule)
                )
            )
        await asyncio.gather(*tasks)

        """Do calcs for rules w values"""
        tasks = []
        for rule in self.rules:
            # TODO uncomment this? maybe not.  save a touch of time. maybe remove and put within function
            if not rule["withValues"]:
                continue
            tasks.append(
                asyncio.create_task(
                    self.do_rule_calculations(rule)
                )
            )
        await asyncio.gather(*tasks)

        """ Write each report to their excel tab """
        for rule in self.rules:
            if rule["formula"]["to"] == "headers":
                header_rules = await self.get_rule_per_header(rule)
                for header_rule in header_rules:
                    await self.write_to_excel(header_rule)
            else:
                await self.write_to_excel(rule)

        #===========================================================
        file_location = None
        """ Save the excel file to the local tmp dir """
        if DEV:
            # LOCAL DEV PURPOSES ONLY
            self.wb.save(filename=self.local_save)
            file_location = self.local_save
        else:
            # PROD (debug or not)
            self.wb.save(filename=self.report_local_path)
            file_location = self.report_local_path

        """ Upload to s3, and return the link """
        self.s3_client.upload_file(
            file_location, self.bucket_name, self.bucket_out_path)
        link = os.environ["bucket_link_path"] + self.bucket_out_path

        """Update job"""
        if (self.job_id != None):
            job_oid = ObjectId(self.job_id)
            jobs_coll = self.mongo.get_collection("jobs")
            job = await jobs_coll.find_one({"_id": job_oid})
            start = job["startDate"]
            end = datetime.now()
            elapsed = end - start
            min, sec = divmod(elapsed.total_seconds(), 60)
            elapsed = f"{int(min)}:{int(sec)}"
            link = {"link" : link}
            up = await jobs_coll.update_one({"_id": job_oid}, {
                "$set": {"endDate": datetime.now(), "isComplete" : True , "info": link, "elapsed" : elapsed},
            })
            print(f"for: {self.mongo.event['job_id']}; matched: {up.matched_count}; mod: {up.modified_count}")
        # print(link)
        return link

    # ====================================================================================================


async def main(event):
    report_generator = ReportGenerator(event)
    link = await report_generator.main()
    return link

if __name__ == '__main__':    
    event = {
        'start_date': '2021/09/12', 
        'end_date': '2021/09/14', 
        # 'rule_ids': ['VISIBLE_EMISSIONS'], 
        'rule_ids': [{ '_id': '61377f37e4a3fb4b98911d24', 'type': 'numeric' }], 
        'flare_id': '5fb6fb02b496f2ae0e0e6845', 
        'debug': False, 
        'timezone': 'America/New_York', 
        'job_id': '60cb6ad151768d63f3f61682'} 


    lambda_handler(event, None)
