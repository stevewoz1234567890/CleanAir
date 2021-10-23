from openpyxl import load_workbook
from datetime import datetime, timedelta, timezone


def get_rows(sheet):
    for row in sheet.iter_rows():
        yield [cell.value for cell in row]


def getHeaderData(sheet):
    for row in sheet.iter_rows(max_row=5):
        yield [cell.value for cell in row]


def getSheetData(sheet, start, end):
    for row in sheet.iter_rows(min_row=start, max_row=end):
        yield [cell.value for cell in row]
        # yield row


filename = '2020-10.xlsx'
date_format = '%Y-%m-%d %H:%M:%S'
workbook = load_workbook(filename=filename,
                         data_only=True, read_only=True)
sheets = workbook.sheetnames
sheet = workbook['15']
rows = get_rows(sheet)
headers = getHeaderData(sheet)
rowNum = 1
piTagIds = []
flareIds = []
PiTags = []
skipNum = 2000
blockCount = 20

dataRows = getSheetData(sheet, skipNum, skipNum+blockCount)
# wb = load_workbook(filename=filename)
# print('hard loaded')
while(True):
    row = next(headers, None)
    if(row == None):
        break
    print(row)


# while(True):
#     row = next(dataRows, None)
#     if(row == None):
#         break
#     print(row)


workbook.close()

# while(True):
#     row = next(rows)
#     if(rowNum == skipNum + blockCount):
#         break
#     if(rowNum == 1):
#         piTagIds = row
#         rowNum = rowNum + 1
#         continue
#     if(rowNum == 2):
#         flareIds = row
#         rowNum = rowNum + 1
#         continue
#     if(rowNum == 5):
#         PiTags = row
#         rowNum = rowNum + 1
#         continue
#     if(rowNum < 6):
#         rowNum = rowNum + 1
#         continue
#     if(rowNum < skipNum):
#         rowNum = rowNum + 1
#         continue

#     dateTime = datetime.strptime(str(row[2]), date_format)
#     for column in range(len(piTagIds)):
#         if(column < 3):
#             continue
#         piTag = piTagIds[column]
#         flare = flareIds[column]
#         value = row[column]
#         print({'dateTime': dateTime, 'piTag': piTag,
#                'flare': flare, 'value': value})

#     rowNum = rowNum + 1
