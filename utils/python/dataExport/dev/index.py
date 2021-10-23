from bson.objectid import ObjectId
import json
from dateutil.parser import parse
import asyncio
from openpyxl import load_workbook, Workbook
import boto3
from AWSUtils import AWSUtils  # Our class
from Mongo import Mongo
import traceback

s3 = boto3.client('s3')

IS_WINDOWS = True  # running on...
# C:\\Users\\evaquero\\Desktop\\data_dump\\    #CAE
# "C:\\Users\\evaquero\\Desktop\\data_dump\\"
WINDOWS_PATH = "C:\\Users\\Erick\\Desktop\\data_dump\\"


class DataExportMain():
    def __init__(self, event):
        self.event = event
        self.event['start'] = parse(self.event['start'], ignoretz=True)
        self.event['end'] = parse(self.event['end'], ignoretz=True)
        self.inputDataKey = event['inputDataKey']


    def build_excel_workbook(self, inputData):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        
        for i in range(65,85):
            ws1.column_dimensions[chr(i)].width = 18
        # ws2 = wb.create_sheet("OIDs")

        # First row is names
        row_data = [outputItem['parameter'] for outputItem in inputData['displayInfo']]
        row_data.insert(0, '')  # PI Tag ID or Formula Name
        ws1.append(row_data)

        # Second row is param name or formula
        row_data = [outputItem['parent'] for outputItem in inputData['displayInfo']]
        row_data.insert(0, '')
        ws1.append(row_data)

        #adding dates and data
        for row in inputData['data']:
            ws1.append(row)
        return wb

    async def save_workbook(self, wb):
        self.file_name = 'data_export_' + str(self.event['start']) + "_" + str(self.event['end']) + '.xlsx'
        self.file_name = self.file_name.replace(" ", "T")
        self.file_name = self.file_name.replace(":", "-")
        self.local_path = '/tmp/' + self.file_name
        if IS_WINDOWS:
            self.local_path = WINDOWS_PATH + self.file_name
            self.local_path = self.local_path.replace("\\\\", "\\")
            print("local path: ", self.local_path)
            wb.save(filename=self.local_path)
            return self.local_path
        else:
            wb.save(filename=self.local_path)
            aws_client = AWSUtils()
            aws_client.upload_file_to_s3(
                bucket='flare-reporting',
                local_file_path=self.local_path,
                s3_file_path='cleancloud-data-export/' + self.file_name
            )
            return 'https://flare-reporting.s3.amazonaws.com/cleancloud-data-export/' + self.file_name

    async def main(self):
        dataFilePath = '/tmp/inputData.json' if not IS_WINDOWS else (WINDOWS_PATH + "inputData.json")
        s3.download_file(
            os.environ["FLARE_REPORTING_BUCKET"], self.inputDataKey, dataFilePath)
        with open(dataFilePath) as f:
            data = json.load(f)
        
        """Build the excel file"""
        wb = self.build_excel_workbook(data['data'])
        save_response = await self.save_workbook(wb)
        return save_response

# class JSONEncoder(json.JSONEncoder):
#     def default(self, o):
#         if isinstance(o, ObjectId):
#             return "hi"#f"ObjectId({str(o)})"
#         return json.JSONEncoder.default(self, o)


def json_print(data, tabs=1):
    print(json.dumps(data, indent=4*tabs, default=str))


async def main(event):
    app = DataExportMain(event)
    url = await app.main()
    return url


def lambda_handler(event, context):
    try:
        print("Event: ", event)
        url = asyncio.run(
            main(event)
        )
        mongo = Mongo(os.environ["DBURI"], os.environ["DB_COLL"])
        response = {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Headers': 'Content-Type, X-Api-Key',
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'OPTIONS,POST,GET'
            },
            'body': url
        }
        mongo.complete_job(ObjectId(event["jobID"]),url)
        mongo.disconnect()
        print("response: ", response)
        return response
    except Exception as e:
        print("Exception: ", e)
        mongo.fail_job(ObjectId(event["jobID"]))
        traceback.print_exc()
        try:
            mongo.disconnect()
        except:
            print("failed to disconnect mongo")


in_event = {
    "org": "5fb6b7ea6b029226f07d2677",
    "tz": "America/New_York",
    "format": "data-dump",
    "debug": True,
    "requested": [
        {
          "id": "5fdd46b8f40caa29a4024774",
          "type": "pitag"
        },
        {
            "id": "5fdd46b8f40caa29a4024773",
            "type": "pitag"
        },
        {
            "id": "5fbd80330aa1ec4824bfc1cf",
            "type": "formula",
            "flare": "5fb6fac8b496f2ae0e0e6844",
            "header": None
        }
    ],
    "end": "2021-04-02T23:30",
    "start": "2021-04-02T11:30",
    "jobID": "606c7f9fbc79ab151026d3e2",
    "inputDataKey": "cleancloud-data-export/606c7f9fbc79ab151026d3e2.json"
}

# the following is useful to make this script executable in both
# AWS Lambda and any other local environments
if __name__ == '__main__':
    import config as os
    lambda_handler(in_event, None)
else:
    import os
